import openpyxl
import re

# open workbook 
excel_file1 = "C:\\Users\\Administrator.DEMO\\Desktop\\svmvol.xlsx"
wb1 = openpyxl.load_workbook(excel_file1)
ws1 = wb1["Sheet1"]

excel_file2 = "C:\\Users\\Administrator.DEMO\\Desktop\\clstrsvm.xlsx"
wb2 = openpyxl.load_workbook(excel_file2)
ws2 = wb2["Sheet1"]

# fn to search all sheets in workbook
def myfind(wb,s):
    for ws in wb.worksheets:       
        for c in range(1,ws.max_column+1):
            for r in range(1,ws.max_row+1):
                txt = ws.cell(r,c).value 
                if txt is None:
                    pass
                elif re.search(s,txt):
                    print("Found",s,txt,ws,r,c)

# scan col C
for r in range(1,ws2.max_row+1):
    s = ws2.cell(r, 3).value 
    if s is None: 
        pass
    else: 
        print(r,s)
        myfind(wb1,s)